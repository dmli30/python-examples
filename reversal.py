#!/usr/bin/env python3
import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string

wb1 = openpyxl.load_workbook('example.xlsx')
wb2 = openpyxl.Workbook()
ws1 = wb1.active
ws2 = wb2.active
for cellOfRows in ws1.rows:
    for cell in cellOfRows:
        ws2[get_column_letter(cell.row) + str(column_index_from_string(cell.column))].value = cell.value
wb1.close()
wb2.save('example.xlsx')
