#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
N=int(input('please input your number: '))
for i in range(2,N+2):    #初始化行1和列A
    sheet.cell(row=i, column=1).value = i - 1
    sheet.cell(row=i, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=i).value = i - 1
    sheet.cell(row=1, column=i).font = Font(bold=True)
for rowNum in range(2,N+2):   #计算填充乘法表
    for columnNum in range(2,N+2):
        sheet.cell(row=rowNum,column=columnNum).value = sheet.cell(row=rowNum,column=1).value * sheet.cell(row=1,column=columnNum).value
filename = str(N) + 'x' + str(N) + '.xlsx'
wb.save(filename)
