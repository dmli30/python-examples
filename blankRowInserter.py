#!/usr/bin/env python3
import openpyxl

def rowCopy (ws1,row1,ws2,row2):    #定义行复制函数，以减少冗余代码
    for cell in ws1[row1]:
        ws2[cell.column + str(row2)].value = cell.value
    return 0

def blankRowInserter (N, M, fileName):    #以行为操作单位，而不是直接遍历所有单元格，避免了在for循环中使用过多判断
    wb1 = openpyxl.load_workbook(fileName)
    wb2 = openpyxl.Workbook()
    ws1 = wb1.active
    ws2 = wb2.active
    if N == 1:    # range(1,1)这种特殊情况单独处理
        for rowNum in range(1,ws1.max_row + 1):
            rowCopy(ws1,rowNum,ws2,rowNum + M)
    elif N > 1 and N <= ws1.max_row:
        for rowNum in range(1,N):
            rowCopy(ws1,rowNum,ws2,rowNum)
        for rowNum in range(N,ws1.max_row + 1):
            rowCopy(ws1,rowNum,ws2,rowNum + M)
    else:
        return 0
    wb1.close()
    wb2.save(fileName)

N = int(input('please input the row number: '))
M = int(input('how many lines would you like to insert: '))
fileName = input('please input the excel file name: ')
blankRowInserter(N, M, fileName)
